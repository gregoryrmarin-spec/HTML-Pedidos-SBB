from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import subprocess
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_HTML = ROOT / "index.html"
DEFAULT_DOWNLOADS = Path.home() / "Downloads"
DATA_RE = re.compile(r"const DATA = (\[.*?\]);\s*\n\s*let pedido", re.S)
STOCK_FILE_RE = re.compile(r"^Saída1 \(\d+\)\.xlsx$", re.I)
LOG_FILE: Path | None = None


def log(message: str) -> None:
    line = f"[{dt.datetime.now():%H:%M:%S}] {message}"
    print(line, flush=True)
    if LOG_FILE:
        LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with LOG_FILE.open("a", encoding="utf-8") as fh:
            fh.write(line + "\n")


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text.strip().upper())


def normalize_isbn(value: object) -> str:
    return re.sub(r"\D", "", "" if value is None else str(value))


def parse_qty(value: object) -> int:
    if value is None:
        return 0
    try:
        qty = int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        qty = 0
    return max(0, qty)


def find_header(ws):
    for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
        headers = [normalize_header(v) for v in row]
        if "ISBN" in headers and "DISPONIBLE" in headers:
            return row_number, headers.index("ISBN"), headers.index("DISPONIBLE")
        if row_number > 30:
            break
    return None


def workbook_has_stock_layout(path: Path) -> bool:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        return any(find_header(ws) for ws in wb.worksheets)
    finally:
        wb.close()


def find_stock_file(downloads: Path) -> Path:
    candidates = sorted(
        (p for p in downloads.glob("*.xlsx") if STOCK_FILE_RE.fullmatch(p.name)),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"Nenhum arquivo 'Saída1 (numero).xlsx' encontrado em {downloads}")

    path = candidates[0]
    if not workbook_has_stock_layout(path):
        raise ValueError(f"O arquivo selecionado pelo nome nao tem colunas ISBN e DISPONIBLE: {path}")
    return path


def load_stock(path: Path) -> tuple[dict[str, int], dict[str, object]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header = find_header(ws)
            if not header:
                continue
            header_row, isbn_col, disponible_col = header
            stock: dict[str, int] = {}
            duplicates = 0
            negative_rows = 0
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                isbn = normalize_isbn(row[isbn_col] if isbn_col < len(row) else None)
                if not isbn:
                    continue
                raw_qty = row[disponible_col] if disponible_col < len(row) else None
                try:
                    qty_before_floor = int(Decimal(str(raw_qty).strip().replace(",", ".")))
                except (InvalidOperation, ValueError):
                    qty_before_floor = 0
                if qty_before_floor < 0:
                    negative_rows += 1
                if isbn in stock:
                    duplicates += 1
                stock[isbn] = max(0, qty_before_floor)
            return stock, {
                "sheet": ws.title,
                "header_row": header_row,
                "isbn_col": isbn_col + 1,
                "disponible_col": disponible_col + 1,
                "duplicates": duplicates,
                "negative_rows": negative_rows,
            }
    finally:
        wb.close()
    raise ValueError(f"Nao encontrei cabecalho ISBN/DISPONIBLE em {path}")


def load_html_data(html_path: Path) -> tuple[str, list[dict]]:
    html = html_path.read_text(encoding="utf-8")
    match = DATA_RE.search(html)
    if not match:
        raise ValueError(f"Nao encontrei 'const DATA = [...]' em {html_path}")
    return html, json.loads(match.group(1))


def update_html(html: str, products: list[dict], generated_at: str) -> str:
    data_json = json.dumps(products, ensure_ascii=False, separators=(",", ":"))
    updated = DATA_RE.sub(f"const DATA = {data_json};\nlet pedido", html, count=1)
    return re.sub(
        r"Generado en \d{2}/\d{2}/\d{4} \d{2}:\d{2}",
        f"Generado en {generated_at}",
        updated,
    )


def run_git(repo: Path, *args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(["git", *args], cwd=repo, text=True, capture_output=True, encoding="utf-8")


def ensure_clean_and_current(repo: Path) -> None:
    status = run_git(repo, "status", "--porcelain")
    if status.returncode != 0:
        raise RuntimeError(status.stderr or status.stdout)
    if status.stdout.strip():
        raise RuntimeError("Repositorio com mudancas locais. Resolva antes de atualizar automaticamente.")

    fetch = run_git(repo, "fetch", "origin", "main")
    if fetch.returncode != 0:
        raise RuntimeError(fetch.stderr or fetch.stdout)
    pull = run_git(repo, "pull", "--ff-only", "origin", "main")
    if pull.returncode != 0:
        raise RuntimeError(pull.stderr or pull.stdout)


def git_publish(repo: Path, html_path: Path) -> bool:
    rel = html_path.relative_to(repo)
    add = run_git(repo, "add", str(rel))
    if add.returncode != 0:
        raise RuntimeError(add.stderr or add.stdout)

    status = run_git(repo, "status", "--porcelain", str(rel))
    if status.returncode != 0:
        raise RuntimeError(status.stderr or status.stdout)
    if not status.stdout.strip():
        log("Nada para publicar no git.")
        return False

    msg = f"Atualiza estoque {dt.datetime.now():%Y-%m-%d %H:%M}"
    commit = run_git(repo, "commit", "-m", msg)
    if commit.returncode != 0:
        raise RuntimeError(commit.stderr or commit.stdout)
    push = run_git(repo, "push", "origin", "main")
    if push.returncode != 0:
        raise RuntimeError(push.stderr or push.stdout)
    log(f"Publicado no git: {msg}")
    return True


def main() -> int:
    global LOG_FILE

    parser = argparse.ArgumentParser(description="Atualiza o estoque disponivel embutido no HTML do catalogo.")
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML, help="Caminho do index.html.")
    parser.add_argument("--stock", type=Path, help="Planilha diaria. Se omitido, usa o arquivo Saída1 (numero).xlsx mais recente em Downloads.")
    parser.add_argument("--downloads", type=Path, default=DEFAULT_DOWNLOADS, help="Pasta onde procurar a planilha diaria.")
    parser.add_argument("--log-dir", type=Path, help="Pasta dos logs. Padrao: logs dentro do repositorio.")
    parser.add_argument("--dry-run", action="store_true", help="Calcula o impacto sem gravar o HTML.")
    parser.add_argument("--push", action="store_true", help="Faz git pull, add, commit e push do index.html.")
    args = parser.parse_args()

    html_path = args.html.resolve()
    repo = html_path.parent
    log_dir = args.log_dir.resolve() if args.log_dir else repo / "logs"
    LOG_FILE = log_dir / f"atualizar_estoque_{dt.datetime.now():%Y-%m-%d}.log"
    log("=" * 72)
    log(f"Log: {LOG_FILE}")
    log("Criterio da planilha diaria: arquivo mais recente em Downloads com nome Saída1 (numero).xlsx")

    if args.push and not args.dry_run:
        log("Sincronizando repositorio antes da atualizacao...")
        ensure_clean_and_current(repo)

    stock_path = (args.stock.resolve() if args.stock else find_stock_file(args.downloads)).resolve()
    generated_at = dt.datetime.now().strftime("%d/%m/%Y %H:%M")

    log(f"HTML: {html_path}")
    log(f"Planilha de estoque: {stock_path}")
    stock, meta = load_stock(stock_path)
    log(
        "Mapeamento: aba {sheet}, ISBN coluna {isbn_col}, DISPONIBLE coluna {disponible_col}, "
        "{negative_rows} linhas negativas convertidas para 0".format(**meta)
    )

    html, products = load_html_data(html_path)
    changed = 0
    matched = 0
    missing = []
    zero_after = 0

    for product in products:
        isbn = normalize_isbn(product.get("e"))
        old_qty = int(product.get("i") or 0)
        if isbn in stock:
            matched += 1
            new_qty = stock[isbn]
        else:
            new_qty = 0
            missing.append(isbn)
        if old_qty != new_qty:
            changed += 1
        if new_qty == 0:
            zero_after += 1
        product["i"] = new_qty

    log(
        f"Catalogo: {len(products)} SKUs, {matched} encontrados na planilha, "
        f"{len(missing)} ausentes tratados como 0, {changed} alteracoes."
    )
    if missing:
        log("SKUs ausentes na planilha diaria: " + ", ".join(missing[:20]))
    log(f"SKUs com disponibilidade 0 apos atualizar: {zero_after}")

    updated_html = update_html(html, products, generated_at)
    if updated_html == html:
        log("HTML ja estava atualizado.")
        return 0

    if args.dry_run:
        log("Dry-run: nenhum arquivo foi gravado.")
        return 0

    html_path.write_text(updated_html, encoding="utf-8")
    log("HTML atualizado.")

    if args.push:
        git_publish(repo, html_path)

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        log(f"ERRO: {exc}")
        raise SystemExit(1)
